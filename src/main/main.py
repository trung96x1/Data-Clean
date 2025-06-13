# Build exe file : pyinstaller --onefile --noconsole .\main_selenium.py
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from deep_translator import GoogleTranslator
from fuzzywuzzy import fuzz
from datetime import datetime
import os
from Aplication import Application as App
from Setting import Setting
from Logger import Logger


from Define import *

table = None
listExporter = []
listImporter = []
setting = None
logger = Logger()

class Table:
    def __init__(self, fileName, sheetName):
        self.workbook = load_workbook(fileName)
        self.worksheet = self.workbook[sheetName]
        self.numcol = self.worksheet.max_column
        self.numrow = self.worksheet.max_row
    
    def addColumnToEnd(self, colName):
        self.worksheet.cell(row=1, column=self.numcol + 1).value = colName
        self.numcol += 1
        columLetter = get_column_letter(self.numcol)
        self.worksheet.column_dimensions[columLetter].width = 20
        return self.numcol
    
    def getCellValue(self, rowIndex, colIndex):
        if rowIndex > self.numrow or colIndex > self.numcol:
            return None
        cell = self.worksheet.cell(row=rowIndex, column=colIndex)
        if cell.value == None:
            return None
        return cell.value
    
    def setCellValue(self, rowIndex, colIndex, value):
        if rowIndex > self.numrow or colIndex > self.numcol:
            return False
        cell = self.worksheet.cell(row=rowIndex, column=colIndex)
        cell.value = value
        return True
    
    def setCellColor(self, rowIndex, colIndex, color):
        if rowIndex > self.numrow or colIndex > self.numcol:
            return False
        cell = self.worksheet.cell(row=rowIndex, column=colIndex)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        return True
    
    def setInvalidCell(self, rowIndex, colIndex):
        if rowIndex > self.numrow or colIndex > self.numcol:
            return False
        self.setCellValue(rowIndex, colIndex, NA)
        self.setCellColor(rowIndex, colIndex, RED_CODE)
        return True
    
    def fillColumColor(self, colIndex, color):
        for row in range(2, self.numrow + 1):
            self.setCellColor(row, colIndex, color)

    def findColumIndex(self, colName):
        for col in range(1, self.numcol + 1):
            if self.worksheet.cell(row=1, column=col).value == colName:
                return col
        return -1

    def save(self, fileName):
        self.workbook.save(fileName)

def setCountry(row, exportCountryIndex, importCountryIndex):
    logger.logi(f"setCountry")
    global table
    if table == None or exportCountryIndex <= 0 or importCountryIndex <= 0 or table.numrow <= 1:
        logger.loge("setCountry : Invalid table or index")
        return False
    
    datasetIndex = table.findColumIndex(DATASET_COLUMN)
    if datasetIndex < 0:
        logger.loge(f"setCountry : {DATASET_COLUMN} not found")
        table.setInvalidCell(row, exportCountryIndex)
        table.setInvalidCell(row, importCountryIndex)
        return False

    dataValue = table.getCellValue(row, datasetIndex)
    if "Export" in dataValue: 
        table.setCellValue(row, exportCountryIndex, dataValue.split("(Export)")[0].strip())

        destCountryIndex = table.findColumIndex(DESTINATION_COUNTRY_COLUMN)
        if destCountryIndex < 0:
            logger.loge(f"setCountry : {DESTINATION_COUNTRY_COLUMN} not found")
            table.setInvalidCell(row, importCountryIndex)
            return False
        destCountry = table.getCellValue(row, destCountryIndex)
        if destCountry == None or destCountry == "":
            logger.loge(f"setCountry : {DESTINATION_COUNTRY_COLUMN} is empty at row {row}")
            table.setInvalidCell(row, importCountryIndex)
        else :
            table.setCellValue(row, importCountryIndex, destCountry)
    elif "Import" in dataValue: 
        table.setCellValue(row, importCountryIndex, dataValue.split("(Import)")[0].strip())

        originCountryIndex = table.findColumIndex(ORIGIN_COUNTRY_COLUMN)
        if originCountryIndex < 0:
            logger.loge(f"setCountry : {ORIGIN_COUNTRY_COLUMN} not found")
            table.setInvalidCell(row, exportCountryIndex)
            return False
        originCountry = table.getCellValue(row, originCountryIndex)
        if originCountry == None or originCountry == "":
            logger.loge(f"setCountry : {ORIGIN_COUNTRY_COLUMN} is empty at row {row}")
            table.setInvalidCell(row, exportCountryIndex)
        else :
            table.setCellValue(row, exportCountryIndex, originCountry)
    else :
        logger.loge(f"setCountry : {DATASET_COLUMN} value is not valid at row {row}")
        table.setInvalidCell(row, exportCountryIndex)
        table.setInvalidCell(row, importCountryIndex)
    return True

def setProduct(row, productIndex):
    logger.logi(f"setProduct")
    global setting 
    global table

    if table == None or productIndex <= 0 or table.numrow <= 1:
        logger.loge("setProduct : Invalid table or index")
        return False
    
    descriptionIndex = table.findColumIndex(DESCRIPTION_COLUMN)
    if descriptionIndex < 0:
        logger.loge(f"setProduct : {DESCRIPTION_COLUMN} not found")
        table.setInvalidCell(row, productIndex)
        return False

    descriptionValue = table.getCellValue(row, descriptionIndex)
    if descriptionValue == None or descriptionValue == "":
        logger.loge(f"setProduct : {DESCRIPTION_COLUMN} is empty at row {row}")
        table.setInvalidCell(row, productIndex)
        return True
    descriptionValue = descriptionValue.lower()
    listProducts = setting.get("listProduct", [])
    for product in listProducts:  
        if any(keyword in descriptionValue for keyword in product.get("key")):
            table.setCellValue(row, productIndex, product.get("name"))
            return True
    
    engTranslated = GoogleTranslator(source='auto', target='en').translate(descriptionValue)
    logger.logi(f"Translated : {engTranslated}")
    for product in listProducts:  
        if any(keyword in descriptionValue for keyword in product.get("key")):
            table.setCellValue(row, productIndex, product.get("name"))
            return True

    logger.loge(f"setProduct : Can't find product for row {row}")
    table.setInvalidCell(row, productIndex)

def setExporter(row, exporterIndex):
    logger.logi(f"setExporter")
    global table, listExporter, setting
    if table == None or exporterIndex <= 0 or table.numrow <= 1:
        logger.loge("setExporter : Invalid table or index")
        return False
    
    listExcludeName = setting.get("listExcludeName", [])
    
    rawExprorterIndex = table.findColumIndex(EXPORTER_COLUMN)
    if rawExprorterIndex < 0:
        logger.loge(f"setExporter : {EXPORTER_COLUMN} not found")
        table.setInvalidCell(row, exporterIndex)
        return False

    rawExporterValue = table.getCellValue(row, rawExprorterIndex)
    rawExporterValue = rawExporterValue.strip().lower()
    for item in listExcludeName:
        rawExporterValue = rawExporterValue.replace(item, "")

    for company in listExporter:
        if fuzz.ratio(company, rawExporterValue) >= 80:
            table.setCellValue(row, exporterIndex, company)
            return True
    listExporter.append(rawExporterValue)
    table.setCellValue(row, exporterIndex, rawExporterValue)

def setImporter(row, importerIndex):
    logger.logi(f"setImporter")
    global table, listImporter, setting
    if table == None or importerIndex <= 0 or table.numrow <= 1:
        logger.loge("setImporter : Invalid table or index")
        return False
    
    listExcludeName = setting.get("listExcludeName", [])
    
    rawImporterIndex = table.findColumIndex(IMPORTER_COLUMN)
    if rawImporterIndex < 0:
        logger.loge(f"setImporter : {IMPORTER_COLUMN} not found")
        table.setInvalidCell(row, importerIndex)
        return False

    rawImporterValue = table.getCellValue(row, rawImporterIndex)
    rawImporterValue = rawImporterValue.strip().lower()
    for item in listExcludeName:
        rawImporterValue = rawImporterValue.replace(item, "")

    for company in listImporter:
        if fuzz.ratio(company, rawImporterValue) >= 70:
            table.setCellValue(row, importerIndex, company)
            return True
    listImporter.append(rawImporterValue)
    table.setCellValue(row, importerIndex, rawImporterValue)

def setUnitPrice(row, unitPriceIndex, quantityIndex, valueIndex):
    logger.logi(f"setUnitPrice")
    global table, setting
    if table == None or unitPriceIndex <= 0 or quantityIndex <= 0 or table.numrow <= 1:
        logger.loge("setUnitPrice : Invalid table or index")
        return False
    
    tempQuantity = -1
    tempUnitPrice = -1
    mainQuantity = -1
    mainUnitPrice = -1
    mainValue = -1

    # Get ratio from Quanity unit
    ratio = -1
    rawQuantityUnitIndex = table.findColumIndex(QUANTITY_UNIT_COLUMN)
    if rawQuantityUnitIndex < 0:
        logger.loge(f"setUnitPrice : {QUANTITY_UNIT_COLUMN} not found")
        ratio = -1
    else :
        rawQuantityUnitValue  = table.getCellValue(row, rawQuantityUnitIndex).strip().lower()
        if rawQuantityUnitValue == None or rawQuantityUnitValue == "":
            logger.loge(f"setUnitPrice : Quantity unit is empty at row {row}")
            ratio = -1
        else:
            weightUnit = setting.get("weightUnit", [])
            for unit in weightUnit:
                unitExchange = unit.get("exchange")
                unitKeys = unit.get("key")
                for key in unitKeys:
                    if rawQuantityUnitValue == key:
                        ratio = unitExchange
                        break
                if ratio != -1:
                    break
    
    # Get temp quantity and temp unit price
    if ratio == -1:
        logger.loge(f"setUnitPrice : Invalid quantity unit at row {row}")
        tempQuantity = -1
        tempUnitPrice = -1
    else:
        rawQuantityIndex = table.findColumIndex(QUANTITY_COLUMN)
        if rawQuantityIndex < 0:
            logger.loge(f"setUnitPrice : {QUANTITY_COLUMN} not found")
            tempQuantity = -1
        else : 
            rawQuantityValue  = table.getCellValue(row, rawQuantityIndex)
            if not isinstance(rawQuantityValue, (int, float)):
                logger.loge(f"setUnitPrice : Invalid data type {QUANTITY_COLUMN} at row {row}")
                tempQuantity = -1
            else :
                tempQuantity = rawQuantityValue * ratio
        
        rawUnitPriceIndex = table.findColumIndex(UNIT_PRICE_COLUMN)
        if rawUnitPriceIndex < 0:
            logger.loge(f"setUnitPrice : {UNIT_PRICE_COLUMN} not found")
            tempUnitPrice = -1
        else :
            rawUnitPriceValue  = table.getCellValue(row, rawUnitPriceIndex)
            if not isinstance(rawUnitPriceValue, (int, float)):
                logger.loge(f"setUnitPrice : Invalid data type {UNIT_PRICE_COLUMN} at row {row}")
                tempUnitPrice = -1
            else :
                tempUnitPrice = rawUnitPriceValue * ratio
    
    # Get weight
    weight = -1
    rawWeightIndex = table.findColumIndex(WEIGHT_COLUMN)
    if rawWeightIndex < 0:
        logger.loge(f"setUnitPrice : {WEIGHT_COLUMN} not found")
        weight = -1
    else :
        rawWeightValue  = table.getCellValue(row, rawWeightIndex)
        if not isinstance(rawWeightValue, (int, float)):
            logger.loge(f"setUnitPrice : Invalid data type {WEIGHT_COLUMN} at row {row}")
            weight = -1
        else :
            weight = rawWeightValue

    # Get value
    value = -1
    rawValueIndex = table.findColumIndex(VALUE_COLUMN)
    if rawValueIndex < 0:
        logger.loge(f"setUnitPrice : {VALUE_COLUMN} not found")
        value = -1
    else :
        rawValueValue  = table.getCellValue(row, rawValueIndex)
        if not isinstance(rawValueValue, (int, float)):
            logger.loge(f"setUnitPrice : Invalid data type {VALUE_COLUMN} at row {row}")
            value = -1
        else :
            value = rawValueValue

    # Get main Quantity
    if tempQuantity == -1 and weight == -1:
        if value != -1 and tempUnitPrice != -1:
            mainQuantity = value / tempUnitPrice
        else:
            mainQuantity = -1
    else:
        # If one of tempQuantity and weight valid or both valid, get the greater value
        mainQuantity = tempQuantity if tempQuantity > weight else weight

    
    # Get main unitPrice
    if tempUnitPrice != -1:
        mainUnitPrice = tempUnitPrice
    else:
        if value != -1 and tempQuantity != -1:
            mainUnitPrice = value / tempQuantity
        else:
            mainUnitPrice = -1

    #Get main value
    if value != -1:
        mainValue = value
    else:
        if tempQuantity != -1 and tempUnitPrice != -1:
            mainValue = tempQuantity * tempUnitPrice
        else:
            mainValue = -1
    
    if mainQuantity != -1:
        table.setCellValue(row, quantityIndex, round(mainQuantity, 2))
    else:
        table.setInvalidCell(row, quantityIndex)

    if mainUnitPrice != -1:
        table.setCellValue(row, unitPriceIndex, round(mainUnitPrice,2))
    else:
        table.setInvalidCell(row, unitPriceIndex)
    
    if mainValue != -1:
        table.setCellValue(row, valueIndex, round(mainValue,2))
    else:
        table.setInvalidCell(row, valueIndex)

def setTime(row, monthIndex, yearIndex):
    logger.logi(f"setTime")
    global table
    if table == None or monthIndex <= 0 or yearIndex <= 0 or table.numrow <= 1:
        logger.loge("setTime : Invalid table or index")
        return False
    
    rawDateIndex = table.findColumIndex(DATE_COLUMN)
    if rawDateIndex < 0:
        logger.loge(f"setTime : {DATE_COLUMN} not found")
        table.setInvalidCell(row, monthIndex)
        table.setInvalidCell(row, yearIndex)
        return False

    rawDateValue = table.getCellValue(row, rawDateIndex)

    date_obj = datetime.strptime(rawDateValue, "%Y-%m-%d")
    if date_obj is None:
        logger.loge(f"setTime : Invalid date format at row {row}")
        table.setInvalidCell(row, monthIndex)
        table.setInvalidCell(row, yearIndex)
        return False

    table.setCellValue(row, monthIndex, date_obj.month)
    table.setCellValue(row, yearIndex, date_obj.year)

class Scenario:
    def __init__(self,):
        pass 
    def execute(self, file_path, app):
        global table
        startTime = datetime.now()
        if file_path:
            table = Table(file_path, SHEET_NAME)
        else:
            table = Table(TEST_FILE, SHEET_NAME)
        exportCountryIndex = table.addColumnToEnd(EXPORT_COUNTRY_COLUMN)
        table.fillColumColor(exportCountryIndex, YELLOW_CODE)  # Yellow

        importCountryIndex = table.addColumnToEnd(IMPORT_COUNTRY_COLUMN)
        table.fillColumColor(importCountryIndex, YELLOW_CODE)  # Yellow

        productIndex = table.addColumnToEnd(PRODUCT_COLUMN)
        table.fillColumColor(productIndex, YELLOW_CODE)  # Yellow

        exportIndex = table.addColumnToEnd(EXPORTER2_COLUMN)
        table.fillColumColor(exportIndex, YELLOW_CODE)  # Yellow

        importIdex = table.addColumnToEnd(IMPORTER2_COLUMN)
        table.fillColumColor(importIdex, YELLOW_CODE)  # Yellow

        unitPriceIndex = table.addColumnToEnd(UNIT_PRICE2_COLUMN)
        table.fillColumColor(unitPriceIndex, YELLOW_CODE)  # Yellow

        quantityIndex = table.addColumnToEnd(QUANTITY_KG_COLUMN)
        table.fillColumColor(quantityIndex, YELLOW_CODE)  # Yellow

        valueIndex = table.addColumnToEnd(VALUE2_COLUMN)
        table.fillColumColor(valueIndex, YELLOW_CODE)

        monthIndex = table.addColumnToEnd(MONTH_COLUMN)
        table.fillColumColor(monthIndex, YELLOW_CODE)  # Yellow

        yearIndex = table.addColumnToEnd(YEAR_COLUMN)
        table.fillColumColor(yearIndex, YELLOW_CODE)  # Yellow

        for row in range(2, table.numrow + 1):
            logger.logi(f"===============Processing {row}/{table.numrow}==============")
            app.setProgress(row, table.numrow)
            #1 Set the country
            setCountry(row, exportCountryIndex, importCountryIndex)

            #2 Set the product
            setProduct(row, productIndex)

            #3 Set the exporter
            setExporter(row, exportIndex)

            #4 Set the importer
            setImporter(row, importIdex)

            #5 Set the unit price, and quantity
            setUnitPrice(row, unitPriceIndex, quantityIndex, valueIndex)

            #6 Set the time
            setTime(row, monthIndex, yearIndex)
  
        
        dataFoler = os.path.join(os.getcwd(), "Data")
        if not os.path.exists(dataFoler):
            os.makedirs(dataFoler)
        fileName = os.path.basename(file_path) if file_path else TEST_FILE
        filePath = os.path.join(dataFoler, f"{fileName.split('.')[0]}_result.xlsx")

        table.save(filePath)
        os.startfile(filePath)
        
        endTime = datetime.now()
        executionTime = (endTime - startTime)
        app.setExecuteTime(executionTime)
        logger.logi(f"Execution completed in {executionTime}")

        
def main():
    global setting
    setting = Setting(SETTING_FILE_PATH)

    scenario = Scenario()
    app = App(scenario)
    app.run()

if __name__ == "__main__":
    main()